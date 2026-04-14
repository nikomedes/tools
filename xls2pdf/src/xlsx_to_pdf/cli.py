from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle


def safe_filename(value: str, max_length: int = 80) -> str:
    value = str(value).strip()
    value = re.sub(r'[\\/:*?"<>|]+', "_", value)
    value = re.sub(r"\s+", "_", value)
    value = value.strip("._")
    return value[:max_length] or "row"


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def get_effective_cell_value(ws, row: int, col: int) -> str:
    cell = ws.cell(row=row, column=col)

    if cell.value is not None:
        return normalize_cell(cell.value)

    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            top_left_value = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value
            return normalize_cell(top_left_value)

    return ""


def find_last_used_column(ws) -> int:
    last_col = 1

    for row, col in ws._cells.keys():
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            last_col = max(last_col, col)

    for merged_range in ws.merged_cells.ranges:
        last_col = max(last_col, merged_range.max_col)

    return last_col


def find_last_used_row(ws) -> int:
    last_row = 1

    for row, col in ws._cells.keys():
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            last_row = max(last_row, row)

    for merged_range in ws.merged_cells.ranges:
        last_row = max(last_row, merged_range.max_row)

    return last_row


def get_header_spans(ws, last_col: int) -> list[tuple[int, int, int, int]]:
    """
    Liefert SPAN-Koordinaten für ReportLab zurück.
    Format: (start_col_0, start_row_0, end_col_0, end_row_0)

    Wir übernehmen nur Merge-Bereiche, die die Header-Zeilen 1-2 betreffen.
    ReportLab ist 0-basiert, Excel 1-basiert.
    """
    spans: list[tuple[int, int, int, int]] = []

    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # Nur relevante Spalten
        if min_col > last_col:
            continue

        max_col = min(max_col, last_col)

        # Nur Header-Bereich Zeile 1-2 berücksichtigen
        if max_row < 1 or min_row > 2:
            continue

        clipped_min_row = max(min_row, 1)
        clipped_max_row = min(max_row, 2)

        # Excel -> ReportLab (0-basiert)
        spans.append((
            min_col - 1,
            clipped_min_row - 1,
            max_col - 1,
            clipped_max_row - 1,
        ))

    return spans


def build_pdf(
    output_file: Path,
    header_rows: list[list[str]],
    data_row: list[str],
    header_spans: list[tuple[int, int, int, int]],
) -> None:
    styles = getSampleStyleSheet()
    style = styles["BodyText"]
    style.fontName = "Helvetica"
    style.fontSize = 7
    style.leading = 8

    raw_data = header_rows + [data_row]
    table_data = [
        [Paragraph(cell if cell else "&nbsp;", style) for cell in row]
        for row in raw_data
    ]

    doc = SimpleDocTemplate(
        str(output_file),
        pagesize=landscape(A3),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    col_count = len(table_data[0]) if table_data else 1
    usable_width = doc.width
    col_widths = [usable_width / col_count] * col_count

    table = Table(table_data, colWidths=col_widths, repeatRows=2)

    commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EEF5FB")),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("ALIGN", (0, 0), (-1, 1), "CENTER"),
    ]

    for span in header_spans:
        start_col, start_row, end_col, end_row = span
        commands.append(("SPAN", (start_col, start_row), (end_col, end_row)))

    table.setStyle(TableStyle(commands))
    doc.build([table])


def split_xlsx(input_xlsx: Path, output_dir: Path, sheet_name: str | None = None) -> None:
    wb = load_workbook(input_xlsx, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    last_row = find_last_used_row(ws)
    last_col = find_last_used_column(ws)

    if last_row < 3:
        raise ValueError("Es werden mindestens 3 Zeilen benötigt.")

    header = [
        [get_effective_cell_value(ws, 1, c) for c in range(1, last_col + 1)],
        [get_effective_cell_value(ws, 2, c) for c in range(1, last_col + 1)],
    ]

    header_spans = get_header_spans(ws, last_col)

    output_dir.mkdir(parents=True, exist_ok=True)

    created = 0

    for r in range(3, last_row + 1):
        row = [get_effective_cell_value(ws, r, c) for c in range(1, last_col + 1)]

        if all(v == "" for v in row):
            continue

        filename = safe_filename(row[0] or f"row_{r}")
        output_file = output_dir / f"{filename}.pdf"

        i = 1
        while output_file.exists():
            output_file = output_dir / f"{filename}_{i}.pdf"
            i += 1

        build_pdf(output_file, header, row, header_spans)
        created += 1

    print(f"Letzte benutzte Spalte: {last_col}")
    print(f"Letzte benutzte Zeile: {last_row}")
    print(f"Header-SPANS: {header_spans}")
    print(f"Fertig. {created} PDF-Datei(en) erstellt.")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_xlsx", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--sheet", default=None)

    args = parser.parse_args()

    try:
        split_xlsx(args.input_xlsx, args.output_dir, args.sheet)
        return 0
    except Exception as e:
        print(f"Fehler: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
